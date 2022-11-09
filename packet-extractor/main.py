import sys
import os
import argparse
import requests
import json
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import datetime
from threading import Thread

import config as conf
from utils.app_path import log_path, source_path, output_folder_path
from utils.app_logger import init_logger, info, error, debug
from utils.app_file_helper import read_lines
from utils.app_session import MosipSession
from utils.app_helper import time_diff, get_time_in_sec, get_timestamp, parse_response
from utils.app_json import dict_to_json
from utils.app_csv import read_csv_file

result_list = []

class ParentThread(Thread):
  def __init__(self, auth_token, rid):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid

  def run(self):
    process_rid(self.rid, self.auth_token)

class IdRepoThread(Thread):
  def __init__(self, auth_token, rid):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.value = None

  def run(self):
    self.value = get_idrepo_identity_by_rid(self.auth_token, self.rid)

class PacketManagerThread(Thread):
  def __init__(self, auth_token, rid):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.value = None

  def run(self):
    self.value =get_info_from_packet(self.auth_token, ["firstName",
         "lastName","middleName","gender","presentProvince"], self.rid, "NEW")

def args_parse():
    parser = argparse.ArgumentParser()
    # group = parser.add_mutually_exclusive_group(required=True)
    # group.add_argument('--print', action='store_true',  help='Create credential for print')
    # group.add_argument('--auth_vid', action='store_true',  help='Create credential for auth(VID)')
    # group.add_argument('--auth_uin', action='store_true',  help='Create credential for auth(UIN)')
    args = parser.parse_args()
    return args, parser

def get_auth_token():
    start_time = get_time_in_sec()
    try:
        regproc_ms = MosipSession(conf.server_url, conf.regproc_client_id, conf.regproc_secret_key, conf.regproc_app_id)
        return regproc_ms.token
    except Exception as e:
        error(f"Exception while getting auth token - {e}")

def main():
    start_time = get_time_in_sec()
    args, parser = args_parse()
    init_logger(log_file=log_path, level=conf.logger_level)
    try:
        if (source_path):
            dict_rows = read_csv_file(source_path)
            if (len(dict_rows) > 0):
                info(f"Number of RIDs to process: {len(dict_rows)}")
                curdate = datetime.today().strftime("%Y%m%d%H%M")
                result_file_name = f"result_{curdate}"
                output_path = os.path.join(output_folder_path, result_file_name + ".xlsx")
                wb = load_workbook("./resource/template.xlsx")
                wb.save(output_path)
                process_rids(dict_rows, output_path)
            else:
                info(f"No RIDs specified in the source file")
        else:
            info(f"Source file not present")
    except Exception as e:
        error(e)
    finally:
        prev_time, prstr = time_diff(start_time)
        info("Total time taken by the script: " + prstr)
        sys.exit(0)

def process_rids(rows, output_path):
    auth_token = get_auth_token()
    count = 0
    row_count_to_write = 0
    thread_list = []
    cur_row_count = 3
    wb = load_workbook(output_path)
    start_time = get_time_in_sec()
    for row in rows:
        rid = row["RID"]
        count = count + 1
        row_count_to_write = row_count_to_write + 1
        thread_list.append(ParentThread(auth_token, rid))
        if (count == conf.thread_count):
            process_thread(thread_list, output_path)
            prev_time, prstr = time_diff(start_time)
            info(f"Time taken to complete {count} packets requests : " + prstr)            
            count = 0
            thread_list = []
            start_time = get_time_in_sec()
        if (row_count_to_write == conf.row_count_to_write):
            row_count_to_write = 0
            cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb)
            result_list.clear()
    if (len(thread_list) > 0):
        process_thread(thread_list, output_path)
        cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb)
        result_list.clear() 
    if (len(result_list) > 0):
        cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb)
        result_list.clear()               

def process_thread(thread_list, output_path):
    for thread in thread_list:
        thread.start()
    for thread in thread_list:
        thread.join()


def write_into_xls(output_path, r_list, cur_row_count, wb):
    s_time = get_time_in_sec()
    ws = wb['Result']
    for result in r_list:
        ws.cell(cur_row_count, 1).value = result["rid"]
        ws.cell(cur_row_count, 2).value = result["status"]
        if ("pktinfo_result" in result):
            ws.cell(cur_row_count, 3).value = json.dumps(result["pktinfo_result"], indent=2)
            ws.cell(cur_row_count, 3).alignment = Alignment(wrap_text=True)
        if ("identity_result" in result):
            ws.cell(cur_row_count, 4).value = json.dumps(result["identity_result"], indent=2)
            ws.cell(cur_row_count, 4).alignment = Alignment(wrap_text=True)
        cur_row_count = cur_row_count + 1
    wb.save(output_path)
    prev_time, prstr = time_diff(s_time)
    info(f"Time taken to complete {len(r_list)} rows to write : " + prstr)     
    return cur_row_count
 

def process_rid(rid, auth_token):
    reg_type="NEW"
    # identity = get_idrepo_identity_by_rid(auth_token, rid)
    # pkt_info = get_info_from_packet(auth_token, ["firstName",
    #     "lastName","middleName","gender","presentProvince"], rid, reg_type)
    t1 = IdRepoThread(auth_token, rid)
    t2 = PacketManagerThread(auth_token, rid)
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    identity = t1.value
    pkt_info = t2.value
    compare(identity, pkt_info, rid)

def compare(identity, pkt_info, rid):
    return_value = {
        "rid": rid
    }
    if (identity and pkt_info):
        identity_firstname = get_value_for(identity["firstName"]) if ("firstName" in identity) else ""
        identity_lastname = get_value_for(identity["lastName"]) if ("lastName" in identity) else ""
        identity_middlename = get_value_for(identity["middleName"]) if ("middleName" in identity) else ""
        identity_gender = get_value_for(identity["gender"]) if ("gender" in identity) else ""
        identity_presentprovince = get_value_for(identity["presentProvince"]) if ("presentProvince" in identity) else ""
        
        pkt_firstname = ""
        if ("firstName" in pkt_info["fields"]):
            pkt_firstname = pkt_info["fields"]["firstName"]
            pkt_firstname = json.loads(pkt_firstname)
            pkt_firstname = get_value_for(pkt_firstname)
        
        pkt_lastname = ""
        if ("lastName" in pkt_info["fields"]):
            pkt_lastname = pkt_info["fields"]["lastName"]
            pkt_lastname = json.loads(pkt_lastname)
            pkt_lastname = get_value_for(pkt_lastname)
        
        pkt_middlename = ""
        if ("middleName" in pkt_info["fields"]):
            pkt_middlename = pkt_info["fields"]["middleName"]
            pkt_middlename = json.loads(pkt_middlename)
            pkt_middlename = get_value_for(pkt_middlename)
        
        pkt_gender = ""
        if ("gender" in pkt_info["fields"]):
            pkt_gender = pkt_info["fields"]["gender"]
            pkt_gender = json.loads(pkt_gender)
            pkt_gender = get_value_for(pkt_gender)
        
        pkt_presentprovince = ""
        if ("presentProvince" in pkt_info["fields"]):
            pkt_presentprovince = pkt_info["fields"]["presentProvince"]
            pkt_presentprovince = json.loads(pkt_presentprovince)
            pkt_presentprovince = get_value_for(pkt_presentprovince)

        identity_result = {
            "firstName": identity_firstname,
            "lastName": identity_lastname,
            "middleName": identity_middlename,
            "gender": identity_gender,
            "province": identity_presentprovince
        }
        pktinfo_result = {
            "firstName": pkt_firstname,
            "lastName": pkt_lastname,
            "middleName": pkt_middlename,
            "gender": pkt_gender,
            "province": pkt_presentprovince
        }    
        return_value["identity_result"] = identity_result
        return_value["pktinfo_result"] = pktinfo_result
        if (identity_firstname != pkt_firstname or
                identity_lastname != pkt_lastname or
                identity_middlename != pkt_middlename or
                identity_gender != pkt_gender or
                identity_presentprovince != pkt_presentprovince):
            return_value["status"] = "Not matching"
            # return_value["identity_result"] = identity_result
            # return_value["pktinfo_result"] = pktinfo_result            
            info(f"RID {rid} - IdRepo and PacketInfo are not matching")
        else:
            return_value["status"] = "Matching"
            info(f"RID {rid} - IdRepo and PacketInfo are matching")
    else:
        status = ""
        if (identity == None):
            info(f"RID {rid} - IdRepo not found")
            status = "IdRepo not found"
        if (pkt_info == None):
            info(f"RID {rid} - Packet Info not found")
            status = status + "\n" + "Packet Info not found" if (len(status) > 0) else "Packet Info not found"
        if (len(status) > 0):
            return_value["status"] = status
    result_list.append(return_value)

def get_value_for(list):
    for d in list:
        if (d.get("language") == "eng"):
            return d.get("value")
    return None

def get_idrepo_identity_by_rid(token, rid):
    try:
        info(f"Identity request api called for {rid}")
        url = '%s/idrepository/v1/identity/idvid/%s' % (conf.server_url, rid)
        params = {
            'idType': 'rid'
        }
        cookies = {'Authorization': token}
        r = requests.get(url, cookies=cookies, verify=conf.ssl_verify, params=params)
        resp = parse_response(r)
        debug("Response: "+ dict_to_json(resp))
        return resp["identity"]
    except Exception as ex:
        error(f"Failed to get identity for RID {rid} - {ex}")
        return None 

def get_info_from_packet(token, field_names_list, rid, reg_type):
    try:
        info(f"Packetmanager searchfield api called for {rid}")
        url = '%s/commons/v1/packetmanager/searchFields' % conf.pkt_mgr_server_url
        cookies = {'Authorization': token}
        ts = get_timestamp()
        req = {
            "id": "packetmanager.searchfield",
            "metadata": {},
            "request": get_searchfield_request(field_names_list, rid, reg_type),
            "requesttime": ts,
            "version": "1.0"
        }
        r = requests.post(url, cookies=cookies, json=req, verify=conf.ssl_verify)
        resp = parse_response(r)
        debug("Response: "+ dict_to_json(resp))
        return resp
    except Exception as ex:
        error(f"Failed to get packet info for RID {rid} - {ex}")
        return None

def get_searchfield_request(field_names_list, rid, reg_type):
    return {
        "bypassCache": False,
        "fields": field_names_list,
        "id": rid,
        "process": reg_type,
        "source": "REGISTRATION_CLIENT"
    }

if __name__ == "__main__":
    main()


